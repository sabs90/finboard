"""
Ingest historical balance sheet data from data/Balance Sheet (1).xlsx.

Loads 30 quarters of data (2018-12-31 to 2026-03-31) into:
  - accounts          (creates balance-sheet accounts on first run)
  - account_balances  (cash + investment quarterly balances)
  - assets            (property valuations)
  - loan_snapshots    (mortgage outstanding + interest rates)
  - net_worth_snapshots (computed per quarter)

Safe to re-run — all writes use upsert logic.

Usage:
    python scripts/ingest_balance_sheet.py [--file path/to/sheet.xlsx] [--dry-run]
"""

import argparse
import datetime
import sys
import time
from pathlib import Path
from typing import Optional

import openpyxl

sys.path.insert(0, str(Path(__file__).parent))

from utils.db import get_connection, transaction
from utils.logger import get_logger

logger = get_logger("ingest_balance_sheet")

XLSX_DEFAULT = Path(__file__).parent.parent / "data" / "Balance Sheet (1).xlsx"


# ── Quarterly date generation ─────────────────────────────────────────────────

def _quarter_end(year: int, month: int) -> datetime.date:
    """Return the last day of the given year/month."""
    if month == 12:
        return datetime.date(year + 1, 1, 1) - datetime.timedelta(days=1)
    return datetime.date(year, month + 1, 1) - datetime.timedelta(days=1)


def generate_quarter_dates() -> list[datetime.date]:
    """
    30 quarter-end dates: 2018-12-31, 2019-03-31, ..., 2026-03-31.
    Index 0 = col 0-based 3 (Excel col D).
    """
    base_year, base_month = 2018, 12
    dates = []
    for i in range(30):
        total_months = (base_year * 12 + base_month - 1) + i * 3
        y, m = divmod(total_months, 12)
        m += 1
        dates.append(_quarter_end(y, m))
    return dates


# ── Row map ───────────────────────────────────────────────────────────────────

CASH_ROWS: list[tuple[int, str, str, str]] = [
    (6,  "ANZ Access (offset)",  "ANZ",   "savings"),
    (7,  "NAB Offset",           "NAB",   "savings"),
    (8,  "Citi",                 "Citi",  "savings"),
    (9,  "ING",                  "ING",   "savings"),
    (19, "AMP",                  "AMP",   "savings"),
    (20, "HSBC - Savings",       "HSBC",  "savings"),
    (21, "HSBC - expense",       "HSBC",  "transaction"),
]

INVESTMENT_ROWS: list[tuple[int, str, str, str]] = [
    (27, "CMC",         "CMC Markets", "investment"),
    (28, "IG Markets",  "IG",          "investment"),
    (29, "Moelis",      "Moelis",      "investment"),
    (30, "Stockland",   "Stockland",   "investment"),
]

OTHER_ASSET_ROWS: list[tuple[int, str, str, str]] = [
    (34, "Crypto",           "Self-managed", "investment"),
    (35, "Qawamah Capital",  "Qawamah",      "investment"),
]

PROPERTY_ROWS: list[tuple[int, str, str, str]] = [
    (49, "8/55 Alice St",  "Property", "property"),
    (50, "8 Yarran St",    "Property", "property"),
]

VEHICLE_ROWS: list[tuple[int, str, str, str]] = [
    (51, "Toyota 86",    "Vehicle", "asset"),
    (52, "Hyundai i30",  "Vehicle", "asset"),
]

LIABILITY_OTHER_ROWS: list[tuple[int, str, str, str]] = [
    (60, "Credit Cards",      "Various", "liability"),
    (61, "Sabeeh - tax bill", "ATO",     "liability"),
]

MORTGAGE_ROWS: list[tuple[int, str, str, str, str]] = [
    (64, "8/55 Alice St (fixed)",    "ANZ", "mortgage", "fixed"),
    (65, "8/55 Alice St (variable)", "ANZ", "mortgage", "variable"),
    (66, "8 Yarran St (fixed)",      "ANZ", "mortgage", "fixed"),
    (67, "8 Yarran St (variable)",   "ANZ", "mortgage", "variable"),
]

# Excel row → account name for interest rate rows
RATE_ROW_MAP: dict[int, str] = {
    113: "8 Yarran St (fixed)",
    117: "8 Yarran St (variable)",
    122: "8/55 Alice St (fixed)",
    126: "8/55 Alice St (variable)",
}


# ── DB helpers ────────────────────────────────────────────────────────────────

def get_or_create_account(conn, name: str, institution: str, account_type: str) -> int:
    row = conn.execute(
        "SELECT id FROM accounts WHERE name = ? AND source = 'manual'", (name,)
    ).fetchone()
    if row:
        return row["id"]
    now = int(time.time())
    conn.execute(
        "INSERT INTO accounts (name, institution, account_type, currency, source, created_at) "
        "VALUES (?, ?, ?, 'AUD', 'manual', ?)",
        (name, institution, account_type, now),
    )
    logger.info("Created account: %s (%s)", name, institution)
    return conn.execute(
        "SELECT id FROM accounts WHERE name = ? AND source = 'manual'", (name,)
    ).fetchone()["id"]


def upsert_account_balance(conn, account_id: int, balance_date: str, balance_cents: int) -> None:
    now = int(time.time())
    conn.execute(
        "INSERT INTO account_balances (account_id, balance_date, balance_cents, source, created_at) "
        "VALUES (?, ?, ?, 'manual', ?) "
        "ON CONFLICT(account_id, balance_date) DO UPDATE SET balance_cents = excluded.balance_cents",
        (account_id, balance_date, balance_cents, now),
    )


def upsert_asset(conn, account_id: int, valuation_date: str, value_cents: int) -> None:
    now = int(time.time())
    existing = conn.execute(
        "SELECT id FROM assets WHERE account_id = ? AND valuation_date = ?",
        (account_id, valuation_date),
    ).fetchone()
    if existing:
        conn.execute(
            "UPDATE assets SET value_cents = ? WHERE id = ?",
            (value_cents, existing["id"]),
        )
    else:
        conn.execute(
            "INSERT INTO assets (account_id, valuation_date, value_cents, source, created_at) "
            "VALUES (?, ?, ?, 'manual', ?)",
            (account_id, valuation_date, value_cents, now),
        )


def upsert_loan_snapshot(
    conn,
    account_id: int,
    snapshot_date: str,
    outstanding_cents: int,
    facility_type: str,
    interest_rate: Optional[float] = None,
) -> None:
    now = int(time.time())
    conn.execute(
        "INSERT INTO loan_snapshots "
        "(account_id, snapshot_date, outstanding_cents, interest_rate, facility_type, source, created_at) "
        "VALUES (?, ?, ?, ?, ?, 'manual', ?) "
        "ON CONFLICT(account_id, snapshot_date) DO UPDATE SET "
        "    outstanding_cents = excluded.outstanding_cents, "
        "    interest_rate = COALESCE(excluded.interest_rate, loan_snapshots.interest_rate), "
        "    facility_type = excluded.facility_type",
        (account_id, snapshot_date, outstanding_cents, interest_rate, facility_type, now),
    )


def upsert_net_worth_snapshot(conn, snap: dict) -> None:
    now = int(time.time())
    conn.execute(
        "INSERT INTO net_worth_snapshots "
        "(snapshot_date, total_assets_cents, total_liabilities_cents, net_worth_cents, "
        " cash_cents, investment_cents, property_value_cents, property_equity_cents, "
        " other_assets_cents, mortgage_cents, other_liabilities_cents, created_at) "
        "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?) "
        "ON CONFLICT(snapshot_date) DO UPDATE SET "
        "    total_assets_cents      = excluded.total_assets_cents, "
        "    total_liabilities_cents = excluded.total_liabilities_cents, "
        "    net_worth_cents         = excluded.net_worth_cents, "
        "    cash_cents              = excluded.cash_cents, "
        "    investment_cents        = excluded.investment_cents, "
        "    property_value_cents    = excluded.property_value_cents, "
        "    property_equity_cents   = excluded.property_equity_cents, "
        "    other_assets_cents      = excluded.other_assets_cents, "
        "    mortgage_cents          = excluded.mortgage_cents, "
        "    other_liabilities_cents = excluded.other_liabilities_cents",
        (
            snap["date"],
            snap["total_assets"], snap["total_liabilities"], snap["net_worth"],
            snap["cash"], snap["investment"], snap["property_value"], snap["property_equity"],
            snap["other_assets"], snap["mortgage"], snap["other_liabilities"],
            now,
        ),
    )


# ── Xlsx reading ──────────────────────────────────────────────────────────────

def to_cents(value) -> Optional[int]:
    """Convert cell value (float/int/None) to integer cents."""
    if value is None:
        return None
    try:
        return round(float(value) * 100)
    except (TypeError, ValueError):
        return None


def read_row_cents(ws, excel_row: int, n: int) -> list[Optional[int]]:
    """Read n quarterly values from excel_row, data starts at col 4 (1-based)."""
    return [to_cents(ws.cell(row=excel_row, column=4 + i).value) for i in range(n)]


def read_row_floats(ws, excel_row: int, n: int) -> list[Optional[float]]:
    """Read n raw float values from a row (for interest rates)."""
    vals = []
    for i in range(n):
        v = ws.cell(row=excel_row, column=4 + i).value
        vals.append(float(v) if v is not None else None)
    return vals


# ── Core ingest ───────────────────────────────────────────────────────────────

def ingest(ws, conn, dry_run: bool) -> None:
    dates = generate_quarter_dates()
    n = len(dates)

    # Build account IDs (skip in dry-run)
    if not dry_run:
        cash_ids    = {name: get_or_create_account(conn, name, inst, atype)
                       for _, name, inst, atype in CASH_ROWS}
        invest_ids  = {name: get_or_create_account(conn, name, inst, atype)
                       for _, name, inst, atype in INVESTMENT_ROWS}
        other_ids   = {name: get_or_create_account(conn, name, inst, atype)
                       for _, name, inst, atype in OTHER_ASSET_ROWS}
        prop_ids    = {name: get_or_create_account(conn, name, inst, atype)
                       for _, name, inst, atype in PROPERTY_ROWS}
        vehicle_ids = {name: get_or_create_account(conn, name, inst, atype)
                       for _, name, inst, atype in VEHICLE_ROWS}
        liab_ids    = {name: get_or_create_account(conn, name, inst, atype)
                       for _, name, inst, atype in LIABILITY_OTHER_ROWS}
        mort_ids    = {name: get_or_create_account(conn, name, inst, atype)
                       for _, name, inst, atype, _ in MORTGAGE_ROWS}

    # Read all row data from xlsx
    cash_vals    = {name: read_row_cents(ws, row, n) for row, name, *_ in CASH_ROWS}
    invest_vals  = {name: read_row_cents(ws, row, n) for row, name, *_ in INVESTMENT_ROWS}
    other_vals   = {name: read_row_cents(ws, row, n) for row, name, *_ in OTHER_ASSET_ROWS}
    prop_vals    = {name: read_row_cents(ws, row, n) for row, name, *_ in PROPERTY_ROWS}
    vehicle_vals = {name: read_row_cents(ws, row, n) for row, name, *_ in VEHICLE_ROWS}
    liab_vals    = {name: read_row_cents(ws, row, n) for row, name, *_ in LIABILITY_OTHER_ROWS}
    mort_vals    = {name: read_row_cents(ws, row, n) for row, name, *_ in MORTGAGE_ROWS}
    mort_types   = {name: ftype for _, name, _, _, ftype in MORTGAGE_ROWS}
    rate_vals    = {acct: read_row_floats(ws, row, n) for row, acct in RATE_ROW_MAP.items()}

    for i, qdate in enumerate(dates):
        ds = qdate.strftime("%Y-%m-%d")

        def c(vals: dict, name: str) -> int:
            return vals[name][i] or 0

        cash_cents      = sum(c(cash_vals, name)    for _, name, *_ in CASH_ROWS)
        invest_cents    = sum(c(invest_vals, name)  for _, name, *_ in INVESTMENT_ROWS)
        other_cents     = sum(c(other_vals, name)   for _, name, *_ in OTHER_ASSET_ROWS)
        prop_cents      = sum(c(prop_vals, name)    for _, name, *_ in PROPERTY_ROWS)
        vehicle_cents   = sum(c(vehicle_vals, name) for _, name, *_ in VEHICLE_ROWS)
        liab_cents      = sum(c(liab_vals, name)    for _, name, *_ in LIABILITY_OTHER_ROWS)
        mort_cents      = sum(c(mort_vals, name)    for _, name, *_ in MORTGAGE_ROWS)

        other_assets    = vehicle_cents + other_cents
        total_assets    = cash_cents + invest_cents + prop_cents + other_assets
        total_liab      = mort_cents + liab_cents
        net_worth       = total_assets - total_liab
        prop_equity     = prop_cents - mort_cents

        if dry_run:
            logger.info(
                "[DRY RUN] %s  NW=$%.0f  Assets=$%.0f  Liabs=$%.0f",
                ds, net_worth / 100, total_assets / 100, total_liab / 100,
            )
            continue

        # account_balances: cash
        for _, name, *_ in CASH_ROWS:
            v = cash_vals[name][i]
            if v is not None:
                upsert_account_balance(conn, cash_ids[name], ds, v)

        # account_balances: investments
        for _, name, *_ in INVESTMENT_ROWS:
            v = invest_vals[name][i]
            if v is not None:
                upsert_account_balance(conn, invest_ids[name], ds, v)

        # account_balances: other (crypto, etc.)
        for _, name, *_ in OTHER_ASSET_ROWS:
            v = other_vals[name][i]
            if v is not None:
                upsert_account_balance(conn, other_ids[name], ds, v)

        # account_balances: vehicles
        for _, name, *_ in VEHICLE_ROWS:
            v = vehicle_vals[name][i]
            if v is not None:
                upsert_account_balance(conn, vehicle_ids[name], ds, v)

        # account_balances: other liabilities
        for _, name, *_ in LIABILITY_OTHER_ROWS:
            v = liab_vals[name][i]
            if v is not None:
                upsert_account_balance(conn, liab_ids[name], ds, v)

        # assets: property valuations
        for _, name, *_ in PROPERTY_ROWS:
            v = prop_vals[name][i]
            if v is not None:
                upsert_asset(conn, prop_ids[name], ds, v)

        # loan_snapshots: mortgages
        for _, name, *_ in MORTGAGE_ROWS:
            v = mort_vals[name][i]
            if v is not None:
                rate = rate_vals.get(name, [None] * n)[i]
                upsert_loan_snapshot(conn, mort_ids[name], ds, v, mort_types[name], rate)

        # net_worth_snapshots
        upsert_net_worth_snapshot(conn, {
            "date":             ds,
            "cash":             cash_cents,
            "investment":       invest_cents,
            "property_value":   prop_cents,
            "property_equity":  prop_equity,
            "other_assets":     other_assets,
            "mortgage":         mort_cents,
            "other_liabilities": liab_cents,
            "total_assets":     total_assets,
            "total_liabilities": total_liab,
            "net_worth":        net_worth,
        })

    if not dry_run:
        logger.info(
            "Processed %d quarters (%s → %s)",
            n, dates[0].strftime("%Y-%m-%d"), dates[-1].strftime("%Y-%m-%d"),
        )


def main() -> None:
    parser = argparse.ArgumentParser(description="Ingest balance sheet xlsx into Finboard")
    parser.add_argument("--file", type=Path, default=XLSX_DEFAULT)
    parser.add_argument("--dry-run", action="store_true", help="Preview without writing to DB")
    args = parser.parse_args()

    if not args.file.exists():
        logger.error("File not found: %s", args.file)
        sys.exit(1)

    logger.info("Loading %s", args.file.name)
    wb = openpyxl.load_workbook(str(args.file), data_only=True)
    ws = wb.active

    conn = get_connection()
    try:
        if args.dry_run:
            ingest(ws, conn, dry_run=True)
        else:
            with transaction(conn):
                ingest(ws, conn, dry_run=False)
    finally:
        conn.close()

    logger.info("Done")


if __name__ == "__main__":
    main()
